import openpyxl
import calendar
import datetime  

def createCronogram(year: int,names: list):
    file_path = "./excel/result.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    month = 1
    day_of_month = 1
    first_day_of_year = datetime.date(year, 1, 1)
    start_day = first_day_of_year.weekday()

    # Rotate the days_of_week array so it starts on the correct day
    days_of_week = ['Lu', 'Ma', 'Mi', 'Ju', 'Vi', 'Sa', 'Do']
    days_of_week = days_of_week[start_day:] + days_of_week[:start_day]

    months = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio',
              'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    num_names = len(names)
    current_week = -1
    for day in range(1, 367):  # year is a leap year, so 366 days
        if day_of_month > calendar.monthrange(year, month)[1]:
            month += 1
            day_of_month = 1
            if month > 12:
                break

        date = datetime.date(year, month, day_of_month)
        day_of_week = days_of_week[date.weekday()]

        cell_day = sheet.cell(row=3, column=day + 1)
        cell_day.value = day_of_month

        cell_weekday = sheet.cell(row=2, column=day + 1)
        cell_weekday.value = day_of_week

        cell_label = sheet.cell(row=2, column=1)
        cell_label.value = "Día"
        
        cell_label = sheet.cell(row=3, column=1)
        cell_label.value = "Nombre"
        
        if day_of_week == 'Ma':
            current_week = (current_week + 1) % num_names

        for i, name in enumerate(names, start=4):
            cell_name = sheet.cell(row=i, column=day+ 1)
            if current_week == i - 4:
                cell_name.value = 'W'
                cell_name.fill = openpyxl.styles.PatternFill(start_color="7033cc", end_color="7033cc", fill_type="solid")  # Set cell color to purple
            else:
                cell_name.value = 'N'
        
        day_of_month += 1
        
    for i, name in enumerate(names, start=4):
        cell_name = sheet.cell(row=i, column=1)
        cell_name.value = name  
    
    sheet.insert_rows(1)
    monthspassed = 0
    for column in sheet.iter_cols(min_col=2, max_col=sheet.max_column):
        if column[3].value == 1:
            column[1].value = months[monthspassed]
            monthspassed += 1
    sheet.delete_rows(1)            
    workbook.save(file_path)
    colorRows()
    
def colorRows():
    file_path = "./excel/result.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    for column in sheet.iter_cols(min_col=2, max_col=sheet.max_column):
        if column[1].value == 'Do' or column[1].value == 'Sa':  # Check if the first cell in the column is 'Do' (Sunday)
            for cell in column:
                if cell.value != 'W':  # Check if the cell value is not 'W'
                    cell.fill = openpyxl.styles.PatternFill(start_color="33cccc", end_color="33cccc", fill_type="solid")  # Set cell color to #33cccc
                    if cell.row >3:
                        cell.value = 'L'
    workbook.save(file_path)
